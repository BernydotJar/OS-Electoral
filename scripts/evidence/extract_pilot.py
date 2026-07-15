#!/usr/bin/env python3
"""Cycle 1 evidence extraction pilot.

Reads original sources from POLITICS_ROOT-relative paths and writes sanitized,
traceable outputs inside research/extracted. Originals are never modified or copied.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
import openpyxl
import pdfplumber
import yaml


EXTRACTOR_NAME = "evidence-pilot-v1"
PII_PATTERNS = [
    (re.compile(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}"), "[REDACTED_EMAIL]"),
    (re.compile(r"(?<!\d)\+?502[\s.-]?\d{4}[\s.-]?\d{4}(?!\d)"), "[REDACTED_PHONE]"),
    (re.compile(r"(?i)\b(?:tel(?:efo(?:no)?)?|telefono|cel(?:ular)?|whatsapp)\s*[:#-]?\s*\d{4}[\s.-]?\d{4}\b"), "[REDACTED_PHONE]"),
    (re.compile(r"(?i)\b(?:dpi|cui)\s*[:#-]?\s*(?:\d[\s-]?){13}\b"), "[REDACTED_DPI]"),
]
PII_NEGATIVE_SELF_TESTS = [
    "CUILAPA",
    "CUILCO",
    "Dirección Municipal de Planificación",
    "Direccion Municipal de Planificacion",
    "10001437",
    "34.293.293",
    "Población 155,383",
]
PII_POSITIVE_SELF_TESTS = [
    "correo persona@example.com",
    "+502 5555 1234",
    "telefono: 5555-1234",
    "DPI 1234567890101",
    "CUI 1234567890101",
]


@dataclass
class ExtractionResult:
    evidence_id: str
    source_type: str
    status: str = "complete"
    outputs: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    page_count: int | None = None
    sheet_names: list[str] = field(default_factory=list)
    empty_pages: list[int] = field(default_factory=list)
    possible_scanned_pages: list[int] = field(default_factory=list)
    table_count: int = 0
    row_count: int = 0
    blank_cell_count: int = 0
    pii_redactions: int = 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract pilot evidence sources.")
    parser.add_argument("--config", required=True, help="YAML config with pilot sources.")
    parser.add_argument("--output", required=True, help="Output directory, normally research/extracted.")
    return parser.parse_args()


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return f"sha256:{digest.hexdigest()}"


def sanitize(value: Any) -> tuple[str, int]:
    if value is None:
        return "", 0
    text = str(value)
    redactions = 0
    for pattern, replacement in PII_PATTERNS:
        text, count = pattern.subn(replacement, text)
        redactions += count
    text = text.replace("\x00", "")
    return text, redactions


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip())
    cleaned = re.sub(r"-+", "-", cleaned).strip("-")
    return cleaned[:80] or "sheet"


def csv_safe_row(writer: csv.writer, row: list[Any]) -> int:
    redactions = 0
    sanitized: list[str] = []
    for value in row:
        text, count = sanitize(value)
        sanitized.append(text)
        redactions += count
    writer.writerow(sanitized)
    return redactions


def resolve_source_path(source: dict[str, Any], politics_root: Path | None = None) -> Path:
    if source.get("resolved_path"):
        return Path(str(source["resolved_path"]))
    if source.get("path"):
        return Path(str(source["path"]))
    relative_path = source.get("relative_path")
    if not relative_path:
        raise ValueError(f"{source.get('evidence_id')} must define path or relative_path")
    if politics_root is None:
        raise ValueError(f"{source.get('evidence_id')} uses relative_path but no politics root is configured")
    return politics_root / str(relative_path)


def load_config(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle) or {}
    politics_root_env = str(data.get("politics_root_env") or "POLITICS_ROOT")
    default_politics_root = data.get("default_politics_root")
    politics_root_value = os.environ.get(politics_root_env) or default_politics_root
    politics_root = Path(str(politics_root_value)).expanduser() if politics_root_value else None
    sources = data.get("sources")
    if not isinstance(sources, list) or not sources:
        raise ValueError("config must contain a non-empty sources list")
    for source in sources:
        source["resolved_path"] = str(resolve_source_path(source, politics_root))
    return sources


def ensure_output_dirs(output: Path) -> None:
    for name in ["manifests", "municipal-core", "territorio", "logs"]:
        (output / name).mkdir(parents=True, exist_ok=True)


def output_dir_for_source(output: Path, source: dict[str, Any]) -> Path:
    collection = source.get("output_collection")
    if collection not in {"municipal-core", "territorio"}:
        raise ValueError(f"{source.get('evidence_id')} has invalid output_collection: {collection}")
    path = output / collection
    path.mkdir(parents=True, exist_ok=True)
    return path


def extract_pdf(source: dict[str, Any], output: Path, extracted_at: str, content_hash: str) -> ExtractionResult:
    evidence_id = source["evidence_id"]
    source_path = resolve_source_path(source)
    target_dir = output_dir_for_source(output, source)
    markdown_path = target_dir / f"{evidence_id}.md"
    tables_path = target_dir / f"{evidence_id}.tables.csv"
    result = ExtractionResult(evidence_id=evidence_id, source_type="pdf")

    doc = fitz.open(source_path)
    result.page_count = doc.page_count

    markdown_parts = [
        f"# {evidence_id} - {source.get('source_title', evidence_id)}",
        "",
        "```yaml",
        f"evidence_id: {evidence_id}",
        f"source_path: {source_path}",
        f"content_hash: {content_hash}",
        f"extracted_at: {extracted_at}",
        f"extractor: {EXTRACTOR_NAME}",
        "```",
        "",
    ]

    with tables_path.open("w", newline="", encoding="utf-8") as csv_handle:
        writer = csv.writer(csv_handle)
        writer.writerow([
            "evidence_id",
            "source_file",
            "page",
            "table_index",
            "row_index",
            "column_index",
            "column_name",
            "cell_value",
        ])

        plumber_doc = pdfplumber.open(source_path)
        try:
            for page_index in range(result.page_count or 0):
                page_number = page_index + 1
                page = doc.load_page(page_index)
                text, text_redactions = sanitize(page.get_text("text"))
                result.pii_redactions += text_redactions

                if not text.strip():
                    result.empty_pages.append(page_number)
                    result.possible_scanned_pages.append(page_number)

                markdown_parts.extend([
                    f"## Pagina {page_number}",
                    "",
                    "### Texto extraido",
                    "",
                    text.strip() or "[NO_EXTRACTABLE_TEXT]",
                    "",
                ])

                try:
                    tables = plumber_doc.pages[page_index].extract_tables() or []
                except Exception as exc:  # pdfplumber table parsing can fail independently.
                    result.errors.append(f"page {page_number} table extraction failed: {exc}")
                    tables = []

                if tables:
                    markdown_parts.extend(["### Tablas detectadas", ""])
                for table_index, table in enumerate(tables, start=1):
                    result.table_count += 1
                    headers = table[0] if table else []
                    markdown_parts.append(
                        f"- Table {table_index}: rows={max(len(table) - 1, 0)} provenance={evidence_id}.tables.csv page={page_number}"
                    )
                    for row_index, row in enumerate(table or [], start=1):
                        for column_index, cell in enumerate(row or [], start=1):
                            header = ""
                            if headers and column_index <= len(headers):
                                header = headers[column_index - 1]
                            result.pii_redactions += csv_safe_row(writer, [
                                evidence_id,
                                source_path.name,
                                page_number,
                                table_index,
                                row_index,
                                column_index,
                                header,
                                cell,
                            ])
                markdown_parts.append("")
        finally:
            plumber_doc.close()
            doc.close()

    if result.empty_pages:
        result.status = "partial"
        result.warnings.append(f"{len(result.empty_pages)} pages had no extractable text and may require OCR")

    if result.errors and result.status == "complete":
        result.status = "partial"

    markdown_path.write_text("\n".join(markdown_parts), encoding="utf-8")
    result.outputs.extend([str(markdown_path), str(tables_path)])
    return result


def extract_xlsx(source: dict[str, Any], output: Path, extracted_at: str, content_hash: str) -> ExtractionResult:
    evidence_id = source["evidence_id"]
    source_path = resolve_source_path(source)
    target_dir = output_dir_for_source(output, source)
    markdown_path = target_dir / f"{evidence_id}.md"
    result = ExtractionResult(evidence_id=evidence_id, source_type="xlsx")

    formula_wb = openpyxl.load_workbook(source_path, data_only=False, read_only=False)
    value_wb = openpyxl.load_workbook(source_path, data_only=True, read_only=False)
    result.sheet_names = list(formula_wb.sheetnames)

    markdown_parts = [
        f"# {evidence_id} - {source.get('source_title', evidence_id)}",
        "",
        "```yaml",
        f"evidence_id: {evidence_id}",
        f"source_path: {source_path}",
        f"content_hash: {content_hash}",
        f"extracted_at: {extracted_at}",
        f"extractor: {EXTRACTOR_NAME}",
        "```",
        "",
        "## Workbook sheets",
        "",
    ]

    for sheet_name in result.sheet_names:
        ws_formula = formula_wb[sheet_name]
        ws_value = value_wb[sheet_name]
        sheet_slug = slugify(sheet_name)
        csv_path = target_dir / f"{evidence_id}.sheet-{sheet_slug}.csv"
        result.outputs.append(str(csv_path))

        max_row = ws_formula.max_row or 0
        max_column = ws_formula.max_column or 0
        headers: dict[int, str] = {}
        for col in range(1, max_column + 1):
            header_value = ws_formula.cell(row=1, column=col).value
            header_text, redactions = sanitize(header_value)
            result.pii_redactions += redactions
            headers[col] = header_text

        with csv_path.open("w", newline="", encoding="utf-8") as csv_handle:
            writer = csv.writer(csv_handle)
            writer.writerow([
                "evidence_id",
                "workbook",
                "sheet",
                "row",
                "column",
                "column_letter",
                "header",
                "value",
                "formula",
                "is_blank",
            ])

            for row in range(1, max_row + 1):
                result.row_count += 1
                for col in range(1, max_column + 1):
                    formula_cell = ws_formula.cell(row=row, column=col)
                    value_cell = ws_value.cell(row=row, column=col)
                    raw_formula = formula_cell.value
                    formula = raw_formula if isinstance(raw_formula, str) and raw_formula.startswith("=") else ""
                    raw_value = value_cell.value if formula else raw_formula
                    is_blank = raw_value is None and not formula
                    if is_blank:
                        result.blank_cell_count += 1
                    result.pii_redactions += csv_safe_row(writer, [
                        evidence_id,
                        source_path.name,
                        sheet_name,
                        row,
                        col,
                        formula_cell.column_letter,
                        headers.get(col, ""),
                        raw_value,
                        formula,
                        "true" if is_blank else "false",
                    ])

        markdown_parts.extend([
            f"### {sheet_name}",
            "",
            f"- CSV: `{csv_path.name}`",
            f"- Rows: {max_row}",
            f"- Columns: {max_column}",
            f"- Provenance columns: evidence_id, workbook, sheet, row, column, column_letter",
            "",
        ])

    formula_wb.close()
    value_wb.close()
    markdown_path.write_text("\n".join(markdown_parts), encoding="utf-8")
    result.outputs.append(str(markdown_path))
    return result


def write_manifest(
    source: dict[str, Any],
    output: Path,
    result: ExtractionResult,
    extracted_at: str,
    content_hash: str,
) -> Path:
    manifest = {
        "evidence_id": source["evidence_id"],
        "source_path": str(resolve_source_path(source)),
        "source_title": source.get("source_title", source["evidence_id"]),
        "source_type": source["source_type"],
        "collection": source["collection"],
        "source_class": source["evidence_class"],
        "review_status_before": source.get("review_status_before", "unknown"),
        "extracted_at": extracted_at,
        "extractor": EXTRACTOR_NAME,
        "source_date": source.get("source_date", "unknown"),
        "territory": source.get("territory", "unknown"),
        "language": source.get("language", "es"),
        "page_count": result.page_count,
        "sheet_names": result.sheet_names,
        "content_hash": content_hash,
        "version_relationship": source.get("version_relationship", "unknown"),
        "contains_personal_data": bool(source.get("contains_personal_data", False)),
        "extraction_status": result.status,
        "limitations": list(source.get("limitations") or []),
        "outputs": [str(Path(path)) for path in result.outputs],
        "empty_pages": result.empty_pages,
        "possible_scanned_pages": result.possible_scanned_pages,
        "table_count": result.table_count,
        "row_count": result.row_count,
        "blank_cell_count": result.blank_cell_count,
        "pii_redactions": result.pii_redactions,
        "warnings": result.warnings,
        "errors": result.errors,
    }

    if result.possible_scanned_pages:
        manifest["limitations"].append("Some pages had no extractable text and may require OCR.")
    if result.errors:
        manifest["limitations"].append("Extractor logged recoverable errors; review logs before promotion.")

    path = output / "manifests" / f"{source['evidence_id']}.yaml"
    with path.open("w", encoding="utf-8") as handle:
        yaml.safe_dump(manifest, handle, sort_keys=False, allow_unicode=True)
    return path


def extract_source(source: dict[str, Any], output: Path, extracted_at: str) -> ExtractionResult:
    source_path = resolve_source_path(source)
    evidence_id = source["evidence_id"]
    source_type = source["source_type"]
    if not source_path.exists():
        return ExtractionResult(
            evidence_id=evidence_id,
            source_type=source_type,
            status="failed",
            errors=[f"source path does not exist: {source_path}"],
        )

    content_hash = sha256_file(source_path)
    try:
        if source_type == "pdf":
            result = extract_pdf(source, output, extracted_at, content_hash)
        elif source_type == "xlsx":
            result = extract_xlsx(source, output, extracted_at, content_hash)
        else:
            result = ExtractionResult(
                evidence_id=evidence_id,
                source_type=source_type,
                status="failed",
                errors=[f"unsupported source_type: {source_type}"],
            )
    except Exception as exc:
        result = ExtractionResult(
            evidence_id=evidence_id,
            source_type=source_type,
            status="failed",
            errors=[f"extraction failed: {exc.__class__.__name__}: {exc}"],
        )

    manifest_path = write_manifest(source, output, result, extracted_at, content_hash)
    result.outputs.append(str(manifest_path))
    return result


def write_run_log(output: Path, results: list[ExtractionResult], extracted_at: str) -> Path:
    stamp = extracted_at.replace("-", "").replace(":", "").replace("Z", "Z")
    log_path = output / "logs" / f"extraction-run-{stamp}.md"
    lines = [
        "# Extraction Run Log",
        "",
        f"- extracted_at: {extracted_at}",
        f"- extractor: {EXTRACTOR_NAME}",
        f"- source_count: {len(results)}",
        "",
        "| Evidence ID | Type | Status | Pages | Sheets | Tables | Rows | Empty Pages | PII Redactions | Errors |",
        "|---|---|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for result in results:
        lines.append(
            "| {id} | {typ} | {status} | {pages} | {sheets} | {tables} | {rows} | {empty} | {redactions} | {errors} |".format(
                id=result.evidence_id,
                typ=result.source_type,
                status=result.status,
                pages=result.page_count if result.page_count is not None else "",
                sheets=len(result.sheet_names),
                tables=result.table_count,
                rows=result.row_count,
                empty=len(result.empty_pages),
                redactions=result.pii_redactions,
                errors=len(result.errors),
            )
        )
    lines.extend(["", "## Details", ""])
    for result in results:
        lines.append(f"### {result.evidence_id}")
        for warning in result.warnings:
            lines.append(f"- warning: {warning}")
        for error in result.errors:
            lines.append(f"- error: {error}")
        if not result.warnings and not result.errors:
            lines.append("- no warnings or errors")
        lines.append("")
    log_path.write_text("\n".join(lines), encoding="utf-8")
    return log_path


def main() -> int:
    args = parse_args()
    output = Path(args.output)
    ensure_output_dirs(output)
    sources = load_config(Path(args.config))
    extracted_at = utc_now()

    results = [extract_source(source, output, extracted_at) for source in sources]
    log_path = write_run_log(output, results, extracted_at)

    for result in results:
        if result.status == "complete":
            print(f"[OK] {result.evidence_id} {result.source_type.upper()} extracted")
        elif result.status == "partial":
            print(f"[PARTIAL] {result.evidence_id} {result.source_type.upper()} extracted with warnings")
        else:
            print(f"[FAILED] {result.evidence_id} {result.source_type.upper()} extraction failed")
    print(f"[DONE] Log written to {log_path}")
    return 1 if any(result.status == "failed" for result in results) else 0


if __name__ == "__main__":
    raise SystemExit(main())
